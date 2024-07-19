import openpyxl

def optimize_menu(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    optimized_wb = openpyxl.Workbook()
    optimized_ws = optimized_wb.active

    optimized_ws.append(['Week', 'Day', 'Menu'])

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            week_number = int(row[0].split()[1])  # Извлечение номера недели из строки "Неделя 1", "Неделя 2" и т.д.
            for i, day in enumerate(['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']):
                menu = row[i + 1] if i + 1 < len(row) else ''  # Проверяем наличие данных для дня
                optimized_ws.append([week_number, day, menu])
        except (ValueError, IndexError):
            continue

    optimized_wb.save('optimized_menu.xlsx')

if __name__ == "__main__":
    filepath = 'menu_new_structure.xlsx'
    optimize_menu(filepath)
    print("Файл optimized_menu.xlsx успешно создан.")
